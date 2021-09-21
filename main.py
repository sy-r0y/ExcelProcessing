

# Python3 project to process Excel spreadsheets
# Uses the openpyxl package .. from pypi
# Use Pip to install various Python packages to do various tasks

import openpyxl as xl
from openpyxl.chart import Reference, BarChart

def modify_sheet(path):
    wb = xl.load_workbook(path)
    sheet = wb['Sheet1']

    # cell1 = sheet['a1']
    # cell2 = sheet.cell(2, 1)
    # The previous 2 lines serve similar purpose, to extract value of a cell in the worksheet

    # print(f'{cell1} value: {cell1.value}')
    # print(f'{cell2} value: {cell2.value}')
    # print(sheet.max_row)  # sheet object has max_row attribute

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 3)
        print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_cell = sheet.cell(row, 4)
        corrected_cell.value = corrected_price

    print(" ")  # Line Break

    for row in range(2, sheet.max_row+1):
        cell = sheet.cell(row, 4)
        print(cell.value)

    correct_cell = sheet.cell(1,4)
    correct_cell.value = 'corrected price'

    # Now let's create a Chart from the corrected price

    chart_values = Reference(sheet,
                                min_row=2, max_row=sheet.max_row,
                                min_col=4, max_col=4)  # create chart values, initialize via constructor

    # Populate the chart with the chart values i.e add the chart values to the chart
    # Then add the chart to the sheet

    chart = BarChart()
    chart.add_data(chart_values)   # Add the values/data to the chart object

    # Add the chart to the sheet
    sheet.add_chart(chart, 'e2')  # Add the chart values to the chart object, then add the chart to the sheet

    # Now save the updates to the workbook/spreadsheet
    wb.save(path)


# Now ... call the function ... send the file-path as argument/parameter

modify_sheet('D:\\Downloads Chrome\\Python-Tutorial-Supplementary-Materials\\transactions.xlsx')












